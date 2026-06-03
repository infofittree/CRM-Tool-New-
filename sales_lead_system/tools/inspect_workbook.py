"""Inspect workbook structure, formulas, and data validations."""

from __future__ import annotations

import json
import sqlite3
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data" / "backup" / "google_sheet_latest_20260601.xlsx"
DATABASE = ROOT / "database" / "sales_leads.db"


def workbook_summary() -> dict:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=False)
    summary = {}
    for ws in wb.worksheets:
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": cell.value})

        validations = []
        for dv in ws.data_validations.dataValidation:
            validations.append(
                {
                    "type": dv.type,
                    "operator": dv.operator,
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                    "allow_blank": bool(dv.allowBlank),
                    "sqref": str(dv.sqref),
                }
            )

        headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        sample_rows = [
            [ws.cell(row, col).value for col in range(1, min(ws.max_column, 12) + 1)]
            for row in range(1, min(ws.max_row, 8) + 1)
        ]
        summary[ws.title] = {
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "headers": headers,
            "validations": validations,
            "formula_count": len(formulas),
            "formulas": formulas[:100],
            "sample_rows": sample_rows,
        }
    return summary


def sqlite_summary() -> dict:
    if not DATABASE.exists():
        return {}
    con = sqlite3.connect(DATABASE)
    try:
        cur = con.cursor()
        tables = [
            row[0]
            for row in cur.execute(
                "select name from sqlite_master where type='table' and name not like 'sqlite_%' order by name"
            )
        ]
        return {
            table: {
                "columns": cur.execute(f"pragma table_info({table})").fetchall(),
                "row_count": cur.execute(f"select count(*) from {table}").fetchone()[0],
            }
            for table in tables
        }
    finally:
        con.close()


def main() -> None:
    print(
        json.dumps(
            {"workbook": workbook_summary(), "sqlite": sqlite_summary()},
            indent=2,
            default=str,
        )
    )


if __name__ == "__main__":
    main()
